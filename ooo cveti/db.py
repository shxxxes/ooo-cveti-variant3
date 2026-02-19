
from __future__ import annotations

import os
import re
import sqlite3
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import openpyxl


DB_FILE = Path(__file__).resolve().parent / "trade.db"
APP_ROOT = Path(__file__).resolve().parent
ASSETS_PRODUCTS_DIR = APP_ROOT / "assets" / "products"
IMPORT_DIR = APP_ROOT / "import_data"


def get_conn() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON;")
    return conn


def init_db_if_needed() -> None:
    """Create schema + import initial data from xlsx if DB doesn't exist yet."""
    if DB_FILE.exists():
        return

    ASSETS_PRODUCTS_DIR.mkdir(parents=True, exist_ok=True)

    with get_conn() as conn:
        _create_schema(conn)
        _import_roles_users(conn, IMPORT_DIR / "user_import.xlsx")
        _import_products(conn, IMPORT_DIR / "products_import.xlsx")
        _import_pickup_points(conn, IMPORT_DIR / "pickup_points_import.xlsx")
        _import_orders(conn, IMPORT_DIR / "orders_import.xlsx")


def _create_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE role (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        );

        CREATE TABLE user (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            surname TEXT NOT NULL,
            name TEXT NOT NULL,
            patronymic TEXT NOT NULL,
            login TEXT NOT NULL UNIQUE,
            password TEXT NOT NULL,
            role_id INTEGER NOT NULL,
            FOREIGN KEY (role_id) REFERENCES role(id)
        );

        CREATE TABLE product (
            article TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            unit TEXT NOT NULL,
            cost REAL NOT NULL CHECK(cost >= 0),
            max_discount INTEGER NOT NULL CHECK(max_discount BETWEEN 0 AND 100),
            manufacturer TEXT NOT NULL,
            supplier TEXT NOT NULL,
            category TEXT NOT NULL,
            discount INTEGER NOT NULL CHECK(discount BETWEEN 0 AND 100),
            quantity INTEGER NOT NULL CHECK(quantity >= 0),
            description TEXT NOT NULL,
            image_path TEXT
        );

        CREATE TABLE pickup_point (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            address TEXT NOT NULL
        );

        CREATE TABLE "order" (
            id INTEGER PRIMARY KEY,
            order_date TEXT NOT NULL,
            delivery_date TEXT NOT NULL,
            pickup_point_id INTEGER NOT NULL,
            client_name TEXT,
            pickup_code INTEGER NOT NULL,
            status TEXT NOT NULL,
            FOREIGN KEY (pickup_point_id) REFERENCES pickup_point(id)
        );

        CREATE TABLE order_product (
            order_id INTEGER NOT NULL,
            product_article TEXT NOT NULL,
            quantity INTEGER NOT NULL CHECK(quantity > 0),
            PRIMARY KEY (order_id, product_article),
            FOREIGN KEY (order_id) REFERENCES "order"(id) ON DELETE CASCADE,
            FOREIGN KEY (product_article) REFERENCES product(article)
        );
        """
    )


def _split_fio(fio: str) -> tuple[str, str, str]:
    parts = [p for p in fio.split() if p.strip()]
    while len(parts) < 3:
        parts.append("")
    return parts[0], parts[1], parts[2]


def _import_roles_users(conn: sqlite3.Connection, xlsx_path: Path) -> None:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Header: (Роль сотрудника, ФИО, Логин, Пароль)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        role_name, fio, login, password = row
        if not (role_name and fio and login and password):
            continue

        role_name_clean = str(role_name).strip()
        conn.execute("INSERT OR IGNORE INTO role(name) VALUES (?)", (role_name_clean,))
        role_id = conn.execute("SELECT id FROM role WHERE name=?", (role_name_clean,)).fetchone()["id"]

        surname, name, patronymic = _split_fio(str(fio).strip())
        conn.execute(
            """
            INSERT INTO user(surname, name, patronymic, login, password, role_id)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (surname, name, patronymic, str(login).strip(), str(password).strip(), role_id),
        )

    # Ensure 'Клиент' exists (for future extension)
    conn.execute("INSERT OR IGNORE INTO role(name) VALUES ('Клиент')")
    conn.execute("INSERT OR IGNORE INTO role(name) VALUES ('Гость')")


def _safe_copy_product_image(filename: str) -> str | None:
    """Copy image from import folder to assets/products, return relative path or None."""
    if not filename:
        return None
    src = IMPORT_DIR / filename
    if not src.exists():
        return None
    dst = ASSETS_PRODUCTS_DIR / filename
    if not dst.exists():
        dst.write_bytes(src.read_bytes())
    rel = os.path.relpath(dst, APP_ROOT)
    return rel


def _import_products(conn: sqlite3.Connection, xlsx_path: Path) -> None:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Columns:
    # Артикул, Наименование, Единица измерения, Стоимость, Макс скидка, Производитель,
    # Поставщик, Категория, Действующая скидка, Кол-во, Описание, Изображение
    for row in ws.iter_rows(min_row=2, values_only=True):
        (
            article,
            name,
            unit,
            cost,
            max_discount,
            manufacturer,
            supplier,
            category,
            discount,
            quantity,
            description,
            image_filename,
        ) = row

        if not article:
            continue

        rel_image = _safe_copy_product_image(str(image_filename).strip() if image_filename else "")
        conn.execute(
            """
            INSERT INTO product(
                article, name, unit, cost, max_discount, manufacturer, supplier, category,
                discount, quantity, description, image_path
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                str(article).strip(),
                str(name).strip(),
                str(unit).strip(),
                float(cost),
                int(max_discount),
                str(manufacturer).strip(),
                str(supplier).strip(),
                str(category).strip(),
                int(discount),
                int(quantity),
                str(description).strip(),
                rel_image,
            ),
        )


def _import_pickup_points(conn: sqlite3.Connection, xlsx_path: Path) -> None:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    for row in ws.iter_rows(min_row=1, values_only=True):
        addr = row[0]
        if addr and str(addr).strip():
            conn.execute("INSERT INTO pickup_point(address) VALUES (?)", (str(addr).strip(),))


_COMPOSITION_RE = re.compile(r"\s*([A-Za-zА-Яа-я0-9]+)\s*,\s*(\d+)\s*")


def _parse_composition(text: str) -> list[tuple[str, int]]:
    # Example: "А112Т4, 2, G843H5, 2"
    items: list[tuple[str, int]] = []
    if not text:
        return items
    parts = [p.strip() for p in str(text).split(",") if p.strip()]
    # pairwise
    for i in range(0, len(parts) - 1, 2):
        art = parts[i]
        try:
            qty = int(parts[i + 1])
        except Exception:
            continue
        items.append((art, qty))
    return items


def _as_iso_date(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, str):
        # try parse dd.mm.yyyy
        v = value.strip()
        for fmt in ("%d.%m.%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(v, fmt).date().isoformat()
            except Exception:
                pass
    # fallback
    return str(value)


def _import_orders(conn: sqlite3.Connection, xlsx_path: Path) -> None:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    # Header columns:
    # Номер заказа, Состав заказа, Дата заказа, Дата доставки, Пункт выдачи,
    # ФИО клиента, Код для получения, Статус заказа
    for row in ws.iter_rows(min_row=2, values_only=True):
        order_id, composition, order_date, delivery_date, pickup_point_id, client_fio, code, status, *_ = row
        if not order_id:
            continue

        conn.execute(
            """
            INSERT INTO "order"(id, order_date, delivery_date, pickup_point_id, client_name, pickup_code, status)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                int(order_id),
                _as_iso_date(order_date),
                _as_iso_date(delivery_date),
                int(pickup_point_id),
                (str(client_fio).strip() if client_fio else None),
                int(code),
                str(status).strip(),
            ),
        )

        for art, qty in _parse_composition(str(composition) if composition else ""):
            # Skip unknown products (shouldn't happen)
            prod = conn.execute("SELECT 1 FROM product WHERE article=?", (art,)).fetchone()
            if not prod:
                continue
            conn.execute(
                "INSERT INTO order_product(order_id, product_article, quantity) VALUES (?, ?, ?)",
                (int(order_id), art, int(qty)),
            )


@dataclass(frozen=True)
class AuthUser:
    id: int
    fio: str
    role: str


def authenticate(login: str, password: str) -> AuthUser | None:
    login = login.strip()
    password = password.strip()
    with get_conn() as conn:
        row = conn.execute(
            """
            SELECT u.id, u.surname, u.name, u.patronymic, r.name AS role_name
            FROM user u
            JOIN role r ON r.id = u.role_id
            WHERE u.login = ? AND u.password = ?
            """,
            (login, password),
        ).fetchone()
        if not row:
            return None
        fio = f"{row['surname']} {row['name']} {row['patronymic']}".strip()
        return AuthUser(id=int(row["id"]), fio=fio, role=str(row["role_name"]))
